import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Entrar na planilha e extrair o CPF do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

drive = webdriver.Chrome()
drive.get('https://consultcpf-devaprender.netlify.app/')
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    # Entra no verificador de cpf
    sleep(5)
    campo_pesquisa = drive.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    # Verificar se está "em dia" ou "atrasado"
    btn_pesquisar = drive.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    btn_pesquisar.click()
    sleep(4)

    status = drive.find_element(By.XPATH, "//span[@id='statusLabel']")
    if status.text == 'em dia':
        # Se estiver em dia, pegar a data do pagamento e o método do pagamento (cartão, boleto)
        data_pagamento = drive.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = drive.find_element(By.XPATH, "//p[@id='paymentMethod']")

        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pag_fechamento = planilha_fechamento['Sheet1']

        # Adicionando uma nova linha na planilha
        pag_fechamento.append([nome, valor, cpf, vencimento, 'em dia',  data_pagamento_limpo, metodo_pagamento_limpo])

        # Salvando as informações
        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        # Caso contrario (se estiver atrasado) coloque o status como pendente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pag_fechamento = planilha_fechamento['Sheet1']

        pag_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')
